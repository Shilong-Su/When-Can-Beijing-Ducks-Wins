import re
import openpyxl
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import time
import random
from selenium import webdriver

def get_one_page(url):
    try:
        head = ['Mozilla/5.0', 'Chrome/78.0.3904.97', 'Safari/537.36']
        headers = {
            'user-agent':head[random.randint(0, 2)]
        }
        driver = webdriver.Chrome('chromedriver.exe')  # 需要下载对应浏览器的驱动并指定位置
        driver.get(url)  # 可查阅selenium使用文档，除了直接获取url对应页面外，还可模拟点击、悬停、下拉、等待等操作，以使得动态页面完成加载。
        response = driver.page_source
        return response
    except RequestException:
        return None

def get_bs_res(selector, html):
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    if res is None:
        return 'NULL'
    elif len(res) == 0:
        return 'NULL'
    else:
        return res[0].string


def parse_one_page(html):
    score_info = {}

    Time = get_bs_res(
        '#wraper > div.content.clearfix > div.part.blk.compare > p > span',
        html)
    score_info['Time'] = Time
    print('【时间】', Time)

    return score_info

def write_bookinfo_excel(book_info, file):
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    sheet_col = ws.max_column
    i = sheet_row
    j = 1
    for key in book_info:
        ws.cell(i+1, j).value = book_info[key]
        j += 1
    done = ws.max_row - sheet_row
    wb.save(file)
    return done

def read_opponent_get_info(src_file, info_file):
    wb = openpyxl.load_workbook(src_file)
    ws = wb.worksheets[0]
    row = ws.max_row
    done = 0
    for i in range(1, row+1):
        src = ws.cell(i, 1).value
        if src is None:
            continue

        html = get_one_page(str(src))
        opponent_info = parse_one_page(html)
        done += write_bookinfo_excel(opponent_info, info_file)
        if done % 1 == 0:
            print('\n', '——————————', done, 'done——————————', '\n')
        time.sleep(random.uniform(0.5, 1.5))
    return done

if __name__ == '__main__':
    res = read_opponent_get_info('opponent_url_all.xlsx', 'time.xlsx')
    print('\n', '——————————', res, 'done——————————', '\n')