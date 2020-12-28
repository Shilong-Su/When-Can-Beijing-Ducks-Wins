import re
import openpyxl
from requests import RequestException
from bs4 import BeautifulSoup
import time
import random
from selenium import webdriver


def get_one_page(url):
    try:
        driver = webdriver.Chrome('chromedriver.exe')
        driver.get(url)
        response = driver.page_source
        return response
    except RequestException:
        return None


def get_request_res(pattern_text, html):
    pattern = re.compile(pattern_text, re.S)
    res = re.findall(pattern, html)
    if len(res) > 0:
        return res[0].split('<', 1)[0][1:]
    else:
        return 'NULL'


def get_bs_res(selector, html):
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    # print(res)
    if str(res) != '[]':
        pattern = re.compile('\((.*?)%', re.S)
        res = str(re.findall(pattern, str(res)))
        res = res.strip('[\'\']') + '%'
    return res


def get_bs2_res(selector, html):
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    if res is None:
        return 'NULL'
    elif len(res) == 0:
        return 'NULL'
    else:
        return res[0].string


def parse_one_page(html):
    score_info = {}

    Time = get_bs2_res(
        '#wraper > div.content.clearfix > div.part.blk.compare > p > span',
        html)
    score_info['Time'] = Time
    print('【时间】', Time)

    x1 = 0
    for a1 in range(0, 100):
        b1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(a1) +
            ') > td:nth-child(1)',
            html)
        if str(b1) != '[]':
            x1 = x1 + 1
    # ——————————————————————————————————————————————————————————————————————————————————————————————————————————————
    cc = get_bs2_res('#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' + str(
        x1) + ') > td:nth-child(1)', html)
    if cc == '北京':

        Rolle1 = '主场'
        score_info['Rolle1'] = Rolle1
        print('【主场】')

        Name1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(1)',
            html)
        score_info['Name1'] = Name1
        print('【队名】', Name1)

        zwei1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(5)', html)
        score_info['zwei1'] = zwei1
        print('【2分中-投】', zwei1)

        drei1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(6)',
            html)
        score_info['drei1'] = drei1
        print('【3分中-投】', drei1)

        Freiwurf1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(7)',
            html)
        score_info['Freiwurf1'] = Freiwurf1
        print('【罚球中-投】', Freiwurf1)

        Angriffen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(8)',
            html)
        score_info['Angriffen1'] = Angriffen1
        print('【进攻篮板】', Angriffen1)

        Abwehr1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(9)',
            html)
        score_info['Abwehr1'] = Abwehr1
        print('【防守篮板】', Abwehr1)

        Unterstützten1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(10)',
            html)
        score_info['Unterstützten1'] = Unterstützten1
        print('【助攻】', Unterstützten1)

        Foul1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(11)',
            html)
        score_info['Foul1'] = Foul1
        print('【犯规】', Foul1)

        Stehlen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(12)',
            html)
        score_info['Stehlen1'] = Stehlen1
        print('【抢断】', Stehlen1)

        Fehler1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(13)',
            html)
        score_info['Fehler1'] = Fehler1
        print('【失误】', Fehler1)

        Blockschuss1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(14)',
            html)
        score_info['Blockschuss1'] = Blockschuss1
        print('【盖帽】', Blockschuss1)

        Dunk1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(15)',
            html)
        score_info['Dunk1'] = Dunk1
        print('【扣篮】', Dunk1)

        Eingedrungen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(16)',
            html)
        score_info['Eingedrungen1'] = Eingedrungen1
        print('【被侵】', Eingedrungen1)

        Fastbreak1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(17)',
            html)
        score_info['Fastbreak1'] = Fastbreak1
        print('【快攻】', Fastbreak1)

        Ergebnis1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(18)',
            html)
        score_info['Ergebnis1'] = Ergebnis1
        print('【得分】', Ergebnis1)

        # ————————————————————————————————————————————————————————————————————————————————————
        x = 0
        for a in range(0, 100):
            b = get_bs_res(
                '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
                str(a) +
                ') > td:nth-child(1)',
                html)
            if str(b) != '[]':
                x = x + 1

        Rolle = '客场'
        score_info['Rolle'] = Rolle
        print('【客场】')

        Name = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(1)',
            html)
        score_info['Name'] = Name
        print('【队名】', Name)

        zwei = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(5)', html)
        score_info['zwei'] = zwei
        print('【2分中-投】', zwei)

        drei = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(6)',
            html)
        score_info['drei'] = drei
        print('【3分中-投】', drei)

        Freiwurf = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(7)',
            html)
        score_info['Freiwurf'] = Freiwurf
        print('【罚球中-投】', Freiwurf)

        Angriffen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(8)',
            html)
        score_info['Angriffen'] = Angriffen
        print('【进攻篮板】', Angriffen)

        Abwehr = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(9)',
            html)
        score_info['Abwehr'] = Abwehr
        print('【防守篮板】', Abwehr)

        Unterstützten = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(10)',
            html)
        score_info['Unterstützten'] = Unterstützten
        print('【助攻】', Unterstützten)

        Foul = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(11)',
            html)
        score_info['Foul'] = Foul
        print('【犯规】', Foul)

        Stehlen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(12)',
            html)
        score_info['Stehlen'] = Stehlen
        print('【抢断】', Stehlen)

        Fehler = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(13)',
            html)
        score_info['Fehler'] = Fehler
        print('【失误】', Fehler)

        Blockschuss = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(14)',
            html)
        score_info['Blockschuss'] = Blockschuss
        print('【盖帽】', Blockschuss)

        Dunk = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(15)',
            html)
        score_info['Dunk'] = Dunk
        print('【扣篮】', Dunk)

        Eingedrungen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(16)',
            html)
        score_info['Eingedrungen'] = Eingedrungen
        print('【被侵】', Eingedrungen)

        Fastbreak = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(17)',
            html)
        score_info['Fastbreak'] = Fastbreak
        print('【快攻】', Fastbreak)

        Ergebnis = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(18)',
            html)
        score_info['Ergebnis'] = Ergebnis
        print('【得分】', Ergebnis)

        # ——————————————————————————————————————————————————————————————————————————————————————————————————————————————
    else:
        x = 0
        for a in range(0, 100):
            b = get_bs_res(
                '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
                str(a) +
                ') > td:nth-child(1)',
                html)
            if str(b) != '[]':
                x = x + 1

        Rolle = '客场'
        score_info['Rolle'] = Rolle
        print('客场')

        Name = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(1)',
            html)
        score_info['Name'] = Name
        print('【队名】', Name)

        zwei = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(5)', html)
        score_info['zwei'] = zwei
        print('【2分中-投】', zwei)

        drei = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(6)',
            html)
        score_info['drei'] = drei
        print('【3分中-投】', drei)

        Freiwurf = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(7)',
            html)
        score_info['Freiwurf'] = Freiwurf
        print('【罚球中-投】', Freiwurf)

        Angriffen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(8)',
            html)
        score_info['Angriffen'] = Angriffen
        print('【进攻篮板】', Angriffen)

        Abwehr = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(9)',
            html)
        score_info['Abwehr'] = Abwehr
        print('【防守篮板】', Abwehr)

        Unterstützten = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(10)',
            html)
        score_info['Unterstützten'] = Unterstützten
        print('【助攻】', Unterstützten)

        Foul = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(11)',
            html)
        score_info['Foul'] = Foul
        print('【犯规】', Foul)

        Stehlen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(12)',
            html)
        score_info['Stehlen'] = Stehlen
        print('【抢断】', Stehlen)

        Fehler = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(13)',
            html)
        score_info['Fehler'] = Fehler
        print('【失误】', Fehler)

        Blockschuss = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(14)',
            html)
        score_info['Blockschuss'] = Blockschuss
        print('【盖帽】', Blockschuss)

        Dunk = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(15)',
            html)
        score_info['Dunk'] = Dunk
        print('【扣篮】', Dunk)

        Eingedrungen = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(16)',
            html)
        score_info['Eingedrungen'] = Eingedrungen
        print('【被侵】', Eingedrungen)

        Fastbreak = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(17)',
            html)
        score_info['Fastbreak'] = Fastbreak
        print('【快攻】', Fastbreak)

        Ergebnis = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part02.blk > div > table > tbody > tr:nth-child(' +
            str(x) +
            ') > td:nth-child(18)',
            html)
        score_info['Ergebnis'] = Ergebnis
        print('【得分】', Ergebnis)

        # ————————————————————————————————————————————————————————————————————————————
        Rolle1 = '主场'
        score_info['Rolle1'] = Rolle1
        print('【主场】')

        Name1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(1)',
            html)
        score_info['Name1'] = Name1
        print('【队名】', Name1)

        zwei1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(5)', html)
        score_info['zwei1'] = zwei1
        print('【2分中-投】', zwei1)

        drei1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(6)',
            html)
        score_info['drei1'] = drei1
        print('【3分中-投】', drei1)

        Freiwurf1 = get_bs_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(7)',
            html)
        score_info['Freiwurf1'] = Freiwurf1
        print('【罚球中-投】', Freiwurf1)

        Angriffen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(8)',
            html)
        score_info['Angriffen1'] = Angriffen1
        print('【进攻篮板】', Angriffen1)

        Abwehr1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(9)',
            html)
        score_info['Abwehr1'] = Abwehr1
        print('【防守篮板】', Abwehr1)

        Unterstützten1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(10)',
            html)
        score_info['Unterstützten1'] = Unterstützten1
        print('【助攻】', Unterstützten1)

        Foul1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(11)',
            html)
        score_info['Foul1'] = Foul1
        print('【犯规】', Foul1)

        Stehlen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(12)',
            html)
        score_info['Stehlen1'] = Stehlen1
        print('【抢断】', Stehlen1)

        Fehler1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(13)',
            html)
        score_info['Fehler1'] = Fehler1
        print('【失误】', Fehler1)

        Blockschuss1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(14)',
            html)
        score_info['Blockschuss1'] = Blockschuss1
        print('【盖帽】', Blockschuss1)

        Dunk1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(15)',
            html)
        score_info['Dunk1'] = Dunk1
        print('【扣篮】', Dunk1)

        Eingedrungen1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(16)',
            html)
        score_info['Eingedrungen1'] = Eingedrungen1
        print('【被侵】', Eingedrungen1)

        Fastbreak1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(17)',
            html)
        score_info['Fastbreak1'] = Fastbreak1
        print('【快攻】', Fastbreak1)

        Ergebnis1 = get_bs2_res(
            '#wraper > div.content.clearfix > div.part.part01.blk > div > table > tbody > tr:nth-child(' +
            str(x1) +
            ') > td:nth-child(18)',
            html)
        score_info['Ergebnis1'] = Ergebnis1
        print('【得分】', Ergebnis1)

    return score_info


def write_opponentinfo_excel(score_info, file):
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    i = sheet_row
    j = 1
    for key in score_info:
        ws.cell(i + 1, j).value = score_info[key]
        j += 1
    done = ws.max_row - sheet_row
    wb.save(file)
    return done


def read_opponent_get_info(src_file, info_file):
    wb = openpyxl.load_workbook(src_file)
    ws = wb.worksheets[0]
    row = ws.max_row
    done = 0
    for i in range(1, row + 1):
        src = ws.cell(i, 1).value
        if src is None:
            continue

        html = get_one_page(str(src))
        opponent_info = parse_one_page(html)
        done += write_opponentinfo_excel(opponent_info, info_file)
        if done % 1 == 0:
            print('\n', '——————————', done, 'done——————————', '\n')
        time.sleep(random.uniform(0.5, 1.5))
    return done


if __name__ == '__main__':
    res = read_opponent_get_info('opponent_url_all.xlsx', 'opponent_info.xlsx')
    print('\n', '——————————', res, 'done——————————', '\n')
