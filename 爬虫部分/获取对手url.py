import re
from selenium import webdriver
import re
import openpyxl
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import time
import random
from selenium.webdriver.support.select import Select


def get_page_src(html, selector):
    if html is not None:
        soup = BeautifulSoup(html, 'lxml')
        res = soup.select(selector)
        # print(type(str(res)))
        if str(res) != '[]':
            pattern = re.compile('href="(.*?)"', re.S)
            src = re.findall(pattern, str(res))
        else:
            src = '[]'
        return src
    else:
        return []



def write_excel_xlsx(items, file):
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    item_num = len(items)
    for i in range(0, item_num):
        ws.cell(sheet_row + i + 1, 1).value = items[i]
    wb.save(file)
    return item_num


url = 'http://cba.sports.sina.com.cn/cba/team/show/10/'

driver = webdriver.Chrome('chromedriver.exe')  # 需要下载对应浏览器的驱动并指定位置
driver.get(url)  # 可查阅selenium使用文档，除了直接获取url对应页面外，还可模拟点击、悬停、下拉、等待等操作，以使得动态页面完成加载。


if __name__ == '__main__':
    for a in range(0, 16):
        js = 'document.querySelector("#league").style="";'  # js去掉上传文件“input”元素的属性，使之可见
        driver.execute_script(js)
        s1 = Select(driver.find_element_by_id('league'))
        s1.select_by_index(a)
        driver.find_element_by_css_selector('#wraper > div.content > div.part.part04.blk.clearfix > div.choose > div > form > input.select_wrap_submit').submit()
        time.sleep(1)
        search_window = driver.current_window_handle
        pageSource = driver.page_source

        one_loop_done = 0
        html = pageSource
        for b in range(1, 100):
            selector = '#team_league > table > tbody > tr:nth-child('+str(b)+') > td:nth-child(2)'
            src = get_page_src(html, selector)
            if src != '[]':
                print(src)
            else:
                break
            row = write_excel_xlsx(src, 'opponent_url_all.xlsx.xlsx')
            one_loop_done += row
        print(one_loop_done, 'done')
        time.sleep(random.uniform(3,5))
